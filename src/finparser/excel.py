"""Excel workbook generation via openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

from finparser.models import ParseResult

# Formatting constants
INDENT_CHARS = 3  # spaces per indent level
HEADER_FONT = Font(bold=True, size=12)
TOTAL_FONT = Font(bold=True)
NUMBER_FORMAT = '#,##0'
NEGATIVE_NUMBER_FORMAT = '#,##0;(#,##0)'


def _sanitize_sheet_name(title: str) -> str:
    """Make a string safe for use as an Excel sheet name (max 31 chars, no special chars)."""
    # Remove characters Excel doesn't allow in sheet names
    for ch in r"[]:*?/\\":
        title = title.replace(ch, "")
    return title[:31]


def write_workbook(result: ParseResult, output_path: Path) -> Path:
    """Write a ParseResult to an Excel workbook.

    Each FinancialStatement becomes its own sheet. Column A has labels (indented),
    columns B+ have period values. Total rows are bolded.

    Args:
        result: Parsed financial data.
        output_path: Where to save the .xlsx file.

    Returns:
        The output path.
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    for stmt in result.statements:
        sheet_name = _sanitize_sheet_name(stmt.title)
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: Statement title
        ws.cell(row=1, column=1, value=stmt.title).font = HEADER_FONT

        # Row 3: Column headers — blank for A, then period names
        ws.cell(row=3, column=1, value="").font = Font(bold=True)
        for col_idx, period in enumerate(stmt.periods, start=2):
            cell = ws.cell(row=3, column=col_idx, value=period)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Row 4+: Line items
        for row_offset, item in enumerate(stmt.line_items):
            row = row_offset + 4
            # Label with indentation
            label = " " * (INDENT_CHARS * item.indent) + item.label
            label_cell = ws.cell(row=row, column=1, value=label)
            if item.is_total:
                label_cell.font = TOTAL_FONT

            # Values for each period
            for col_idx, period in enumerate(stmt.periods, start=2):
                value = item.values.get(period)
                if value is not None:
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    if isinstance(value, (int, float)):
                        cell.number_format = NEGATIVE_NUMBER_FORMAT
                    if item.is_total:
                        cell.font = TOTAL_FONT

        # Auto-size column A (labels)
        max_label_len = max(
            (len(" " * (INDENT_CHARS * item.indent) + item.label) for item in stmt.line_items),
            default=20,
        )
        ws.column_dimensions["A"].width = min(max_label_len + 2, 60)

        # Size period columns
        for col_idx in range(2, len(stmt.periods) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # If no statements were extracted, keep a blank sheet with a message
    if not result.statements:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No financial statements were extracted.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
